import os
import sys
import openpyxl


def copy_worksheet(ws, new_ws):
    '''Destructively copy a worksheet to a new one'''
    for row in ws.rows:
        for cell in row:
            new_cell = new_ws[cell.coordinate]
            new_cell.value = cell.value
            new_cell.style = cell.style
            new_cell.border = cell.border.copy()
            new_cell.font = cell.font.copy()
            new_cell.fill = cell.fill.copy()
            new_cell.alignment = cell.alignment.copy()
            new_cell.number_format = cell.number_format

    for merged_range in ws.merged_cell_ranges:
        new_ws.merge_cells(merged_range)
        first = merged_range.split(':')[0]
        orig, merged = ws[first], new_ws[first]
        merged.value = orig.value
        cell = ws[first]
        for row in new_ws[merged_range]:
            for new_cell in row:
                new_cell.style = cell.style
                new_cell.border = cell.border.copy()
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format

    for row_id, row_dim in ws.row_dimensions.items():
        new_row = new_ws.row_dimensions[row_id]
        new_row.height = row_dim.height

    for col_id, col_dim in ws.column_dimensions.items():
        new_column = new_ws.column_dimensions[col_id]
        new_column.width = col_dim.width


def combine(spreadsheet_path,
            new_wb=None,
            grep_values=None,
            rack_profiles=True,
            include_filename_in_sheetname=False,
            set_print_area=True):

    if new_wb is None:
        new_wb = openpyxl.Workbook()
        del new_wb['Sheet']

    matched_sheets = []

    for fn in os.listdir(spreadsheet_path):
        try:
            wb = openpyxl.load_workbook(os.path.join(spreadsheet_path, fn))
        except Exception as ex:
            print('Failed to load: {}'.format(fn), ex)
            continue

        for ws in wb.worksheets:
            print('File: {} worksheet: {}'.format(fn, ws))
            if grep_values is not None:
                found = any(to_grep in str(cell.value)
                            for row in ws.rows
                            for cell in row
                            if cell.value is not None
                            for to_grep in grep_values
                            )

                if not found:
                    continue

                print('Match')

            if include_filename_in_sheetname:
                title = '{}_{}'.format(fn.rsplit('.', 1)[0], ws.title)
            else:
                title = ws.title

            if rack_profiles:
                c2 = ws['C2'].value
                if c2 is not None and c2.strip().lower() == 'rack no.':
                    if ws['E2'].value.strip():
                        title = ws['E2'].value.strip()
                        print('New sheet title based on rack number found:',
                              title)

            matched_sheets.append((fn, ws.title, title))
            new_ws = new_wb.create_sheet(title)
            copy_worksheet(ws, new_ws)

            if set_print_area:
                new_ws.print_area = new_ws.dimensions
                new_ws.print_options.horizontalCentered = True
                new_ws.print_options.verticalCentered = True
                header_text = 'File: {} Sheet: {}'.format(fn.rsplit('.', 1)[0],
                                                          ws.title)
                new_ws.firstHeader.center.text = header_text
                new_ws.page_setup.fitToPage = True

    return matched_sheets, new_wb


if __name__ == '__main__':
    matched_sheets, new_wb = combine(spreadsheet_path='copied',
                                     grep_values=('FWS', 'Lauer', 'Bong',
                                                  'AEROTECH', 'Aerotech',
                                                  'LVDT', 'Epaq'),
                                     rack_profiles=True,
                                     include_filename_in_sheetname=False)
    try:
        output_fn = sys.argv[2]
    except IndexError:
        output_fn = 'combined.xlsx'

    new_wb.save(output_fn)
    print('Saved to', output_fn)

    print('Matching sheets: ', len(matched_sheets))
    for fn, title, new_title in matched_sheets:
        print('\t', fn, ':', title, '->', new_title)
